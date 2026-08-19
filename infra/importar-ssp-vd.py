#!/usr/bin/env python3
"""
importar-ssp-vd.py — Violência doméstica em SC (SSP/SC)

FONTE: https://ssp.sc.gov.br/segurancaemnumeros/
       "Indicadores de Violência Doméstica" (XLSX)

FORMATO REAL DA PLANILHA (verificado, nao suposto):
    linha 1: ['Município', 'Fato Comunicado', '2025', ...]
    linha 2: [None, None, 'jan', 'fev', 'mar', 'abr', 'mai', 'jun']
    linha 3: ['ABDON BATISTA', 'Ameaça', 0,0,0,0,1,1]
    linha 4: [None,            'Injúria', 0,0,1,1,0,0]   <-- municipio VAZIO

    O municipio so aparece na primeira linha do grupo (celula mesclada).
    Precisa ser arrastado para baixo, senao "Injúria" fica sem municipio.

IMPORTANTE — POR QUE ISSO NAO VAI PARA O MAPA:
    O dado e MENSAL e AGREGADO por municipio. Nao existe coordenada nem
    data de ocorrencia. Colocar um alfinete no mapa a partir de
    "Palhoca: 12 ameacas em maio" seria inventar uma localizacao que o
    dado nao tem. Por isso guardamos em tabela propria, para uso como
    camada de contexto e comparativo — nao como ponto no mapa.

Uso:
    python3 importar-ssp-vd.py /tmp/vd.xlsx [--gravar]
    (sem --gravar apenas mostra o que faria)
"""
import sys, unicodedata, subprocess, json

# Os 22 municipios da Grande Florianopolis (cobertura do produto)
GRANDE_FLORIANOPOLIS = {
 'FLORIANOPOLIS','SAO JOSE','PALHOCA','BIGUACU','SANTO AMARO DA IMPERATRIZ',
 'AGUAS MORNAS','SAO PEDRO DE ALCANTARA','ANTONIO CARLOS',
 'GOVERNADOR CELSO RAMOS','TIJUCAS','CANELINHA','SAO JOAO BATISTA',
 'MAJOR GERCINO','ANGELINA','RANCHO QUEIMADO','ALFREDO WAGNER',
 'ANITAPOLIS','SAO BONIFACIO','PAULO LOPES','GAROPABA','IMBITUBA','IMARUI'
}

MESES = {'jan':1,'fev':2,'mar':3,'abr':4,'mai':5,'jun':6,
         'jul':7,'ago':8,'set':9,'out':10,'nov':11,'dez':12}

def sem_acento(s):
    if not s: return ''
    n = unicodedata.normalize('NFD', str(s))
    return ''.join(c for c in n if unicodedata.category(c) != 'Mn').upper().strip()

def ler(caminho):
    import openpyxl
    w = openpyxl.load_workbook(caminho, data_only=True)
    s = w[w.sheetnames[0]]
    linhas = list(s.iter_rows(values_only=True))

    # ano fica na linha 1, coluna 3
    ano = None
    for c in linhas[0]:
        if c and str(c).strip().isdigit() and len(str(c).strip()) == 4:
            ano = int(str(c).strip()); break
    if not ano: ano = 2025

    # linha 2 traz os meses; monta indice -> numero do mes
    col_mes = {}
    for i, c in enumerate(linhas[1]):
        chave = str(c).strip().lower()[:3] if c else ''
        if chave in MESES: col_mes[i] = MESES[chave]

    registros = []
    municipio_atual = None   # arrasta o municipio das celulas mescladas

    for linha in linhas[2:]:
        if not linha or all(c is None for c in linha): continue

        bruto_muni = linha[0]
        if bruto_muni and str(bruto_muni).strip():
            municipio_atual = str(bruto_muni).strip()

        fato = linha[1]
        if not municipio_atual or not fato: continue
        if sem_acento(municipio_atual) in ('TOTAL','TOTAL GERAL'): continue

        for idx, mes in col_mes.items():
            if idx >= len(linha): continue
            v = linha[idx]
            try: qtd = int(float(str(v).replace('.','').replace(',','.') or 0))
            except (TypeError, ValueError): continue
            if qtd <= 0: continue   # zero nao e informacao util aqui
            registros.append({
                'municipio': municipio_atual.strip(),
                'municipio_norm': sem_acento(municipio_atual),
                'fato': str(fato).strip(),
                'ano': ano, 'mes': mes, 'total': qtd,
            })
    return registros, ano

def sql_escape(s):
    return str(s).replace("'", "''")

def main():
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(1)
    caminho = sys.argv[1]
    gravar  = '--gravar' in sys.argv

    registros, ano = ler(caminho)
    na_regiao = [r for r in registros if r['municipio_norm'] in GRANDE_FLORIANOPOLIS]

    print(f'Ano de referencia .............. {ano}')
    print(f'Registros lidos ................ {len(registros)}')
    print(f'Na Grande Florianopolis ........ {len(na_regiao)}')
    print(f'Municipios distintos (regiao) .. {len(set(r["municipio"] for r in na_regiao))}')
    print(f'Total de ocorrencias (regiao) .. {sum(r["total"] for r in na_regiao)}')

    print('\nPor municipio (regiao):')
    porm = {}
    for r in na_regiao: porm[r['municipio']] = porm.get(r['municipio'], 0) + r['total']
    for m, t in sorted(porm.items(), key=lambda x: -x[1]):
        print(f'   {m:32} {t:>6}')

    print('\nPor tipo de fato (regiao):')
    porf = {}
    for r in na_regiao: porf[r['fato']] = porf.get(r['fato'], 0) + r['total']
    for f, t in sorted(porf.items(), key=lambda x: -x[1])[:12]:
        print(f'   {f:42} {t:>6}')

    if not gravar:
        print('\n(simulacao — use --gravar para inserir no banco)')
        return

    # grava so a Grande Florianopolis: e a area coberta pelo produto
    linhas_sql = []
    for r in na_regiao:
        linhas_sql.append(
            "('{}','{}',{},{},{},'ssp-sc')".format(
                sql_escape(r['municipio']), sql_escape(r['fato']),
                r['ano'], r['mes'], r['total']))

    sql = ("DELETE FROM ssp_indicadores WHERE ano = {} AND fonte = 'ssp-sc';\n"
           "INSERT INTO ssp_indicadores (municipio, fato, ano, mes, total, fonte) VALUES\n"
           .format(ano) + ',\n'.join(linhas_sql) + ';')

    open('/tmp/ssp-insert.sql','w',encoding='utf-8').write(sql)
    r = subprocess.run(['docker','exec','-i','comunidade-alerta-db-1','psql','-U','postgres',
                        '-d','comunidade_alerta','-f','/dev/stdin'],
                       input=sql, capture_output=True, text=True)
    print('\n' + (r.stdout or '') + (r.stderr or ''))
    print(f'{len(linhas_sql)} registro(s) enviado(s).')

if __name__ == '__main__':
    main()
